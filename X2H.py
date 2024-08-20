

import xlwings as xw
import openpyxl

def read_excel(excel_file_name):
    
    feature_type = []


    # Specifying a sheet
    print(excel_file_name)
    ws = xw.Book(excel_file_name + '.xlsx').sheets['Sheet1']
    print(ws)
    
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(excel_file_name + '.xlsx')

    sheet = dataframe.worksheets[0]

    row_count = sheet.max_row
    column_count = sheet.max_column
    print(row_count)
    print(column_count)

    # Selecting data from
    # a single cell
    # v1 = ws.range("B1:B28").value
    # v2 = ws.range("C1:C28").value
    # v3 = ws.range("D1:D28").value
    # v4 = ws.range("E1:E28").value
    v1 = ws.range("A1:A"+str(int(row_count)+1)).value
    v2 = ws.range("B1:B"+str(int(row_count)+1)).value
    v3 = ws.range("C1:C"+str(int(row_count)+1)).value
    v4 = ws.range("D1:D"+str(int(row_count)+1)).value

    res = {}

    for key in v1:
        for value in v2:
            res[key] = value
            v2.remove(value)
            break

    for i, key in enumerate(res):
        res[key] = (res[key], v3[i], v4[i])

    # print(res)

    # for key, value in res.items():
        # print(key + ' ' + value[0] + ' ' + value[1] + ' ' + value[2] + '\n')

    print("H2X COMP")

    return res



            
